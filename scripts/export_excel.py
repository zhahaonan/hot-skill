#!/usr/bin/env python3
"""
export_excel — Generate Excel report with trend overview, creative briefs, and materials.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from _common import (
    base_argparser, handle_schema, read_json_input, write_json_output,
    fail, today_str,
    format_material_item, material_category_label,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    fail("openpyxl not installed. Run: pip install openpyxl")

SCHEMA = {
    "name": "export_excel",
    "description": "Generate Excel report: 趋势总览 + 创作简报 + 素材库 + 平台标题矩阵 (4 sheets). Also accepts trend_analyze output (briefs sheet will be empty).",
    "input": {
        "type": "object",
        "properties": {
            "briefed_trends": {"type": "array", "description": "Output from content_brief (or trend_analyze with empty briefs)"}
        },
        "required": ["briefed_trends"]
    },
    "output": {
        "type": "object",
        "properties": {
            "file": {"type": "string", "description": "Absolute path to generated .xlsx file"}
        }
    },
    "examples": {
        "cli": "python scripts/export_excel.py -i briefs.json --xlsx output/report.xlsx"
    },
    "errors": {
        "no_data": "无 briefed_trends → 先运行 content_brief",
        "write_error": "文件写入失败 → 检查路径权限"
    }
}

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HOT_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
EMERGING_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def style_header(ws, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, col_count, max_width=40):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 4


def build_overview_sheet(wb, trends):
    ws = wb.active
    ws.title = "趋势总览"
    headers = ["排名", "话题", "热度分", "趋势方向", "类别", "覆盖平台", "一句话概要"]
    ws.append(headers)
    style_header(ws, len(headers))

    for i, t in enumerate(trends, 1):
        platforms_str = ", ".join(t.get("platforms", []))
        direction_map = {"rising": "↑上升", "peak": "●顶峰", "declining": "↓下降", "emerging": "★萌芽"}
        direction = direction_map.get(t.get("direction", ""), t.get("direction", ""))
        row = [
            i,
            t.get("topic", ""),
            t.get("score", 0),
            direction,
            t.get("category", ""),
            platforms_str,
            t.get("summary", "")
        ]
        ws.append(row)

        row_idx = i + 1
        score = t.get("score", 0)
        is_emerging = t.get("is_emerging", False) or t.get("direction") == "emerging"

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if score >= 70:
                cell.fill = HOT_FILL
            elif is_emerging:
                cell.fill = EMERGING_FILL

    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def build_brief_sheet(wb, trends):
    ws = wb.create_sheet("创作简报")
    headers = ["话题", "热度", "洞察", "创作角度", "热词/标签", "标题建议", "推荐形式", "最佳时间"]
    ws.append(headers)
    style_header(ws, len(headers))

    for t in trends:
        brief = t.get("brief", t.get("content_brief", {}))
        if isinstance(brief, dict) and "error" in brief:
            ws.append([t.get("topic", ""), t.get("score", 0), f"错误: {brief['error']}", "", "", "", "", ""])
            continue

        insight = brief.get("insight", "")
        if isinstance(insight, dict):
            insight = f"{insight.get('core', '')} {insight.get('why_hot', '')} {insight.get('opportunity', '')}"

        angles = brief.get("angles", [])
        angles_str = "\n".join(
            f"• {a.get('name', '')}: {a.get('description', '') or a.get('how', '')} [{a.get('best_platform', '')}]"
            for a in angles
        ) if angles else ""

        keywords = brief.get("hot_keywords", [])
        keywords_str = ", ".join(keywords) if keywords else ""
        if not keywords_str:
            tags = t.get("content_brief", {}).get("tags", [])
            keywords_str = ", ".join(tags) if tags else ""

        titles = brief.get("titles", {})
        titles_str = flatten_titles(titles)

        rec = brief.get("recommendation", {})
        best_format = rec.get("best_format", "")
        best_time = rec.get("best_time", "")

        row = [
            t.get("topic", ""),
            t.get("score", 0),
            insight,
            angles_str,
            keywords_str,
            titles_str,
            best_format,
            best_time,
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def build_materials_sheet(wb, trends):
    ws = wb.create_sheet("素材库")
    headers = ["话题", "热度", "关键素材", "对标案例", "原始链接"]
    ws.append(headers)
    style_header(ws, len(headers))

    for t in trends:
        brief = t.get("brief", {})
        if isinstance(brief, dict) and "error" in brief:
            continue

        materials = brief.get("materials", [])
        materials_str = flatten_materials(materials)

        benchmarks = brief.get("benchmarks", [])
        bench_lines = []
        for b in benchmarks:
            who = b.get("author_type", "") or b.get("brand", "") or b.get("creator_type", "")
            why = b.get("reason", "") or b.get("why_viral", "")
            bench_lines.append(f"• [{b.get('platform', '')}] {who} — {b.get('metrics', '')} ({why})")
        bench_str = "\n".join(bench_lines)

        platforms = t.get("platforms", [])
        links = ", ".join(platforms)

        ws.append([
            t.get("topic", ""),
            t.get("score", 0),
            materials_str,
            bench_str,
            links,
        ])

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


PLATFORM_KEYS = ["douyin", "xiaohongshu", "gongzhonghao", "zhihu", "bilibili"]
PLATFORM_NAMES = {"douyin": "抖音", "xiaohongshu": "小红书",
                  "gongzhonghao": "公众号", "zhihu": "知乎", "bilibili": "B站"}


SHORTFORM_LABEL = {"short_form": "短视频/小红书", "long_form": "公众号/知乎"}

def build_titles_sheet(wb, trends):
    """Per-platform title matrix: one row per topic x title."""
    ws = wb.create_sheet("平台标题")
    headers = ["话题", "类别", "平台", "标题"]
    ws.append(headers)
    style_header(ws, len(headers))

    for t in trends:
        brief = t.get("brief", t.get("content_brief", {}))
        if isinstance(brief, dict) and "error" in brief:
            continue
        titles = brief.get("titles", {})
        if not titles:
            continue
        topic = t.get("topic", "")
        category = t.get("category", "")

        has_platform_keys = any(k in titles for k in PLATFORM_KEYS)
        if has_platform_keys:
            for pkey in PLATFORM_KEYS:
                vals = titles.get(pkey, [])
                if isinstance(vals, str):
                    vals = [vals]
                pname = PLATFORM_NAMES.get(pkey, pkey)
                for title in vals:
                    if title:
                        ws.append([topic, category, pname, title])
        else:
            for key, val in titles.items():
                label = SHORTFORM_LABEL.get(key, key)
                if isinstance(val, list):
                    for title in val:
                        if title:
                            ws.append([topic, category, label, title])
                elif val:
                    ws.append([topic, category, label, val])

    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    auto_width(ws, len(headers))
    ws.freeze_panes = "A2"


def flatten_materials(materials) -> str:
    """Handle list or dict; dict values may be strings or structured objects."""
    if isinstance(materials, list):
        return "\n".join(
            f"• {format_material_item(m)}" for m in materials if format_material_item(m)
        )
    if isinstance(materials, dict):
        lines = []
        for category, items in materials.items():
            label = material_category_label(category)
            if isinstance(items, list):
                for item in items:
                    line = format_material_item(item)
                    if line:
                        lines.append(f"• [{label}] {line}")
            else:
                line = format_material_item(items)
                if line:
                    lines.append(f"• [{label}] {line}")
        return "\n".join(lines)
    return str(materials) if materials else ""


def flatten_titles(titles) -> str:
    """Handle both old (dict of str) and new (dict of list) titles format."""
    if not titles or not isinstance(titles, dict):
        return ""
    lines = []
    platform_names = {
        "douyin": "抖音", "xiaohongshu": "小红书",
        "gongzhonghao": "公众号", "zhihu": "知乎", "bilibili": "B站"
    }
    for key, val in titles.items():
        pname = platform_names.get(key, key)
        if isinstance(val, list):
            for t in val:
                lines.append(f"[{pname}] {t}")
        else:
            lines.append(f"[{pname}] {val}")
    return "\n".join(lines)


def main():
    parser = base_argparser("Generate Excel trend report")
    parser.add_argument("--xlsx", help="Output .xlsx file path (default: hot-creator-{date}.xlsx)")
    args = parser.parse_args()
    handle_schema(args, SCHEMA)

    input_data = read_json_input(args)
    trends = input_data.get("briefed_trends", [])

    if not trends:
        fail("No briefed_trends provided. Pipe output from content_brief.")

    output_path = args.xlsx or args.output or f"hot-creator-{today_str()}.xlsx"
    if not output_path.endswith(".xlsx"):
        output_path += ".xlsx"

    wb = Workbook()
    build_overview_sheet(wb, trends)
    build_brief_sheet(wb, trends)
    build_materials_sheet(wb, trends)
    build_titles_sheet(wb, trends)

    from pathlib import Path
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    print(f"[export_excel] Saved to {output_path}", file=sys.stderr)

    import json
    print(json.dumps({"file": str(output_path)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
